import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment
import json

# Load JSON data
with open('data.json') as json_file:
    data = json.load(json_file)

# Create a workbook and a worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Set column widths
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 50
ws.column_dimensions['D'].width = 10

# Merge cells for the header and set the title
ws.merge_cells('C1:D1')
ws['C1'] = 'Project report'
ws['C1'].font = Font(size=24, bold=True)
ws['C1'].alignment = Alignment(horizontal='left', vertical='center')

# Inserting the image
img = Image(data["logo_url"])
print(data["logo_url"])  # Replace 'path_to_your_image.png' with the actual path to your image file
ws.add_image(img, 'E1')

# Populate data from JSON
ws['C3'] = 'Month'
ws['D3'] = data['month']

ws['C4'] = 'Contract ID'
ws['D4'] = data['contract_id']

ws['C5'] = 'Supplier Number'
ws['D5'] = data['supplier_number']

ws['C6'] = 'Consultant Name'
ws['D6'] = data['consultant_name']

# Add the address row
ws.merge_cells('C8:D8')
ws['C8'] = data['address']

# Header for the table
header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
ws['B10'] = 'Days'
ws['B10'].fill = header_fill
ws['C10'] = 'Tasks'
ws['C10'].fill = header_fill
ws['D10'] = 'Hours'
ws['D10'].fill = header_fill

# Populate date-wise tasks
row = 11
for task in data['date_wise_tasks']:
    ws[f'B{row}'] = task['date']
    for sub_task in task['tasks']:
        ws[f'C{row}'] = sub_task['name']
        ws[f'D{row}'] = sub_task['hours']
        row += 1

# Adding the confirmation area
ws.merge_cells('B43:D43')
ws['B43'] = f'Confirmed / Date: {data["confirmation_date"]}'

ws.merge_cells('B44:D44')
ws['B44'] = f'Name (Please Block Letters): {data["display_name"]}'

# Saving the workbook
wb.save('project_report.xlsx')
