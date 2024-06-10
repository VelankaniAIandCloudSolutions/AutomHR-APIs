from flask import Flask, request, jsonify,  send_from_directory
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
import json
import requests
import io
import calendar
from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.styles import Border, Side
import os
from flask import send_file
from datetime import datetime
from spire.pdf.common import *
from spire.pdf import *
from spire.xls import Workbook as SpireWorkbook, FileFormat
from spire.xls.common import *
import fitz  # PyMuPDF
import logging
from logging.handlers import RotatingFileHandler
from dotenv import load_dotenv
import boto3


load_dotenv()
s3_client = boto3.client(
    's3',
    aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
    aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY')
)
bucket_name = os.getenv('AWS_STORAGE_BUCKET_NAME')


def upload_to_s3(file_path, s3_key):
    try:
        s3_client.upload_file(file_path, bucket_name, s3_key)
        s3_url = f"https://{bucket_name}.s3.amazonaws.com/{s3_key}"
        print(f"Uploaded file to S3. URL: {s3_url}")
        return s3_url
    except Exception as e:
        error_msg = f"Error uploading file to S3: {str(e)}"
        print(error_msg)
        logger.error("Error uploading file to S3: %s", str(e))
        raise


app = Flask(__name__)

# Set up logging
log_file_path = os.path.join(os.getcwd(), 'app.log')
handler = RotatingFileHandler(log_file_path, maxBytes=100000, backupCount=3)
handler.setLevel(logging.ERROR)
formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

# Get the logger for your application
logger = logging.getLogger(__name__)
logger.setLevel(logging.ERROR)
logger.addHandler(handler)


def excel_to_pdf(excel_file_path, pdf_file_path):
    workbook = SpireWorkbook()
    workbook.LoadFromFile(excel_file_path)

    workbook.ConverterSetting.SheetFitToPage = True
    print(pdf_file_path)
    workbook.SaveToFile(pdf_file_path, FileFormat.PDF)
    workbook.Dispose()

    doc = PdfDocument()
    doc.LoadFromFile(pdf_file_path)

    for i in range(doc.Pages.Count):
        page = doc.Pages[i]
        replacer = PdfTextReplacer(page)
        replacer.ReplaceAllText(
            "Evaluation Warning : The document was created with Spire.XLS for Python", "")
    doc.SaveToFile(pdf_file_path)
    doc.Close()

    pdf_document = fitz.open(pdf_file_path)
    first_page = pdf_document.load_page(0)
    page_width = first_page.rect.width
    page_height = first_page.rect.height

    fill_height_percentage = 4

    fill_height = min(page_height, int(
        page_height * fill_height_percentage / 100))
    fill_rect = fitz.Rect(0, 0, page_width, fill_height)
    first_page.draw_rect(fill_rect, fill=(1, 1, 1), color=(1, 1, 1))
    new_pdf_file_path = pdf_file_path.replace(".pdf", "_filled.pdf")
    pdf_document.delete_page(1)
    pdf_document.save(new_pdf_file_path)
    pdf_document.close()

    os.remove(pdf_file_path)
    return new_pdf_file_path


@app.route('/api/v1/generate_timesheet_report', methods=['POST'])
def generate_timesheet_report():
    try:
        data = request.json
        header_fill = PatternFill(
            start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

        thick_black_border = Border(left=Side(style='thick', color='000000'),
                                    right=Side(style='thick', color='000000'),
                                    top=Side(style='thick', color='000000'),
                                    bottom=Side(style='thick', color='000000'))

        light_black_border = Border(left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000'))
        # Define border styles
        # Thick border style
        thick_side = Side(border_style="thick", color="000000")
        # Thick border style
        thin_side = Side(border_style="thin", color="000000")
        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active
        image_url = data["logo_url"]

        # Download the image using requests
        response = requests.get(image_url)
        image_data = response.content
        img = Image(io.BytesIO(image_data))
        ws.add_image(img, 'D2')
        ws.row_dimensions[3].height = 27.75

        # Merge cells for the header and set the title
        ws.merge_cells('B4:C4')
        ws['B4'] = 'Project Report ' + \
            data['project_name'] if data['project_name'] else None
        ws['B4'].font = Font(size=22, bold=True, name='Arial')
        ws['B4'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[4].height = 27.75

        # Set column widths
        ws.column_dimensions['A'].width = 0.86
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 72
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 10

        # Populate data from JSON
        ws['B7'] = 'Month'
        ws['B7'].font = Font(bold=True, name='Arial', size=14)
        ws['D7'] = data['month']
        ws['D7'].font = Font(bold=True, color="FF0000", name='Arial', size=14)
        ws['D7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D7'].fill = header_fill

        ws['B8'] = 'Contract ID'
        ws['B8'].font = Font(bold=True, name='Arial', size=14)
        ws['D8'] = data['contract_id']
        ws['D8'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D8'].font = Font(bold=True, name='Arial', size=14)
        ws['D8'].fill = header_fill

        ws['B9'] = 'Supplier Number'
        ws['B9'].font = Font(bold=True, name='Arial', size=14)
        ws['D9'] = data['supplier_number']
        ws['D9'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D9'].font = Font(bold=True, name='Arial', size=14)
        ws['D9'].fill = header_fill

        ws['B10'] = 'Consultant Name'
        ws['B10'].font = Font(bold=True, name='Arial', size=14)
        ws['D10'] = data['consultant_name']
        ws['D10'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D10'].font = Font(bold=True, name='Arial', size=14)
        ws['D10'].fill = header_fill

        thin_red_bottom_border = Border(
            bottom=Side(style='thin', color='000000'))

    # Apply border to specified cells
        for row in range(7, 11):
            ws[f'D{row}'].border = thin_red_bottom_border
        # Add the address row
        ws.merge_cells('B12:C12')
        ws['B12'] = data['address']
        ws['B12'].font = Font(bold=True, name='Arial', size=10)
        ws.row_dimensions[12].height = 34.5

        # Header for the table
        ws['B18'] = 'Days'
        ws['B18'].font = Font(bold=True, name='Arial', size=10)
        ws['B18'].fill = header_fill
        ws['B18'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C18'] = 'Tasks'
        ws['C18'].font = Font(bold=True, name='Arial', size=10)
        ws['C18'].fill = header_fill
        ws['C18'].alignment = Alignment(horizontal='center', vertical='center')
        ws['D18'] = 'Hours'
        ws['D18'].font = Font(bold=True, name='Arial', size=10)
        ws['D18'].fill = header_fill
        ws['D18'].alignment = Alignment(horizontal='center', vertical='center')

        # if 'total_days_worked' in data and data['total_days_worked'] != 0:
        #     ws['E18'] = 'Days Worked'
        #     ws['E18'].font = Font(bold=True, name='Arial', size=10)
        #     ws['E18'].fill = header_fill

        # Populate date-wise tasks
        row = 19

        # Parse the month from the request data (assuming the month format is 'mm/yy')
        weekend_fill_color = "E7E6E6"

        month, year = map(int, data['month'].split('/'))
        num_days_in_month = calendar.monthrange(year, month)[1]
        for day in range(1, num_days_in_month + 1):
            tasks_for_day = None
            for task in data['date_wise_tasks']:
                task_day, task_month, task_year = map(
                    int, task['date'].split('/'))  # Parsing task date
                if task_day == day and task_month == month and task_year == year:  # Comparing with current day
                    tasks_for_day = task
                    break

            # Determine if it's a weekend
            weekday = calendar.weekday(year, month, day)
            # Saturday and Sunday are represented by 5 and 6 respectively in calendar module
            is_weekend = weekday >= 5

            if tasks_for_day:
                # Fill in data for the day if there are tasks
                ws[f'B{row}'] = day
                ws[f'B{row}'].font = Font(bold=True, name='Arial', size=10)
                ws[f'B{row}'].alignment = Alignment(
                    horizontal='center', vertical='center')

                for sub_task in tasks_for_day['tasks']:
                    cell = ws[f'C{row}']
                    cell_value = cell.value + ', ' + \
                        sub_task['name'] + ' ' + sub_task['description'] if cell.value else sub_task['name'] + \
                        ' ' + sub_task['description']
                    # Enable word wrapping for the cell
                    cell.alignment = Alignment(wrapText=True)
                    cell.value = cell_value

                    # Calculate the required row height based on wrapped text and column width
                    # Approximate height per line
                    cell_wrap_height = (
                        len(cell_value) // (ws.column_dimensions['C'].width // 7)) * 2
                    if len(cell_value) % (ws.column_dimensions['C'].width // 7):
                        cell_wrap_height += 15  # Add extra height if there's more content

                    current_row_height = ws.row_dimensions[row].height
                    if current_row_height is None or cell_wrap_height > current_row_height:
                        ws.row_dimensions[row].height = cell_wrap_height

                ws[f'D{row}'] = tasks_for_day['total_hours']
                ws[f'D{row}'].font = Font(bold=True, name='Arial', size=10)
                ws[f'D{row}'].alignment = Alignment(
                    horizontal='center', vertical='center')

                # if 'total_days_worked' in data and data['total_days_worked'] != 0:
                #     ws[f'E{row}'] = tasks_for_day['days_worked']
                #     ws[f'E{row}'].alignment = Alignment(
                #         horizontal='center', vertical='center')

            else:
                # Fill in the day even if there are no tasks
                ws[f'B{row}'] = day
                ws[f'B{row}'].font = Font(bold=True, name='Arial', size=10)
                ws[f'B{row}'].alignment = Alignment(
                    horizontal='center', vertical='center')
                ws[f'D{row}'] = None  # Leave hours blank
                # if 'total_days_worked' in data and data['total_days_worked'] != 0:
                #     ws[f'E{row}'] = 0
                #     ws[f'E{row}'].alignment = Alignment(
                #         horizontal='center', vertical='center')

            # Set background color for weekends with no tasks
            if not tasks_for_day and is_weekend:
                for col in ['B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = PatternFill(
                        start_color=weekend_fill_color, end_color=weekend_fill_color, fill_type="solid")
                # if 'total_days_worked' in data and data['total_days_worked'] != 0:
                #     ws[f'E{row}'].fill = PatternFill(
                #         start_color=weekend_fill_color, end_color=weekend_fill_color, fill_type="solid")

            row += 1

        if data['hour_display'] != 0:
            ws[f'C{row+2}'] = 'Total Hours (in h)'
            ws[f'C{row+2}'].alignment = Alignment(
                horizontal='right', vertical='center')
            ws[f'C{row+2}'].font = Font(bold=True, name='Arial', size=11)

            ws[f'D{row+2}'] = data['total_hours']
            ws[f'D{row+2}'].font = Font(bold=True, name='Arial', size=11)
            ws[f'D{row+2}'].alignment = Alignment(
                horizontal='center', vertical='center')

            ws[f'D{row+2}'].fill = header_fill
            ws[f'D{row+2}'].border = thick_black_border


        if data['day_display'] != 0 and 'total_working_days' in data:
            ws[f'C{row+3}'] = 'Total Days'
            ws[f'C{row+3}'].alignment = Alignment(horizontal='right', vertical='center')
            ws[f'C{row+3}'].font = Font(bold=True, name='Arial', size=11)
            ws[f'D{row+3}'] = data['total_working_days']
            ws[f'D{row+3}'].font = Font(bold=True, name='Arial', size=11)
            ws[f'D{row+3}'].alignment = Alignment(
                horizontal='center', vertical='center')
            ws[f'D{row+3}'].fill = header_fill
            ws[f'D{row+3}'].border = thick_black_border

        # if 'total_days_worked' in data and data['total_days_worked'] != 0:

        #     ws[f'E{row+2}'] = data['total_days_worked']
        #     ws[f'E{row+2}'].font = Font(bold=True, name='Arial', size=11)
        #     ws[f'E{row+2}'].alignment = Alignment(
        #         horizontal='center', vertical='center')

        #     ws[f'E{row+2}'].fill = header_fill
        #     ws[f'E{row+2}'].border = thick_black_border

        # Adding the confirmation area
        ws[f'C{row+5}'] = f'Confirmed / Date: {data["confirmation_date"]}'
        ws[f'C{row+5}'].font = Font(name='Arial', size=10)

        if (data["display_value"] == 'N.A.'):
            ws[f'C{row+6}'] = f'Name ({data["display_label"]})'
            ws[f'C{row+6}'].font = Font(bold=True, name='Arial', size=10)
        else:
            ws[f'C{row+6}'] = f'Name ({data["display_label"]}): {data["display_value"]}'
            ws[f'C{row+6}'].font = Font(bold=True, name='Arial', size=10)

        # if 'total_days_worked' in data and data['total_days_worked'] != 0:
        #     header_cells = ['B18', 'C18', 'D18', 'E18']
        # else:
            # header_cells = ['B18', 'C18', 'D18']
        header_cells = ['B18', 'C18', 'D18']

        for cell_ref in header_cells:
            ws[cell_ref].border = thick_black_border

        for row_item in range(19, row + 1):
            ws[f'B{row_item}'].border = Border(
                left=thick_side, bottom=thin_side, right=thin_side, top=thin_side)
            ws[f'C{row_item}'].border = Border(
                left=thin_side, bottom=thin_side, right=thin_side, top=thin_side)
            ws[f'D{row_item}'].border = Border(
                left=thin_side, bottom=thin_side, right=thick_side, top=thin_side)
            # if 'total_days_worked' in data and data['total_days_worked'] != 0:
            #     ws[f'D{row_item}'].border = Border(
            #         left=thin_side, bottom=thin_side, right=thin_side, top=thin_side)
            #     ws[f'E{row_item}'].border = Border(
            #         left=thin_side, bottom=thin_side, right=thick_side, top=thin_side)

            if row_item == row:
                ws[f'B{row_item}'].border = Border(
                    left=thick_side, bottom=thick_side, right=thin_side, top=thin_side)
                ws[f'C{row_item}'].border = Border(
                    left=thin_side, bottom=thick_side, right=thin_side, top=thin_side)
                ws[f'D{row_item}'].border = Border(
                    left=thin_side, bottom=thick_side, right=thick_side, top=thin_side)
                # if 'total_days_worked' in data and data['total_days_worked'] != 0:
                #     ws[f'D{row_item}'].border = Border(
                #         left=thin_side, bottom=thick_side, right=thin_side, top=thin_side)
                #     ws[f'E{row_item}'].border = Border(
                #         left=thin_side, bottom=thick_side, right=thick_side, top=thin_side)

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        consultant_name = data['consultant_name'].replace(" ", "_")
        unique_filename = f"timesheet_{timestamp}"

        excel_folder_path = os.path.join(os.getcwd(), 'static')
        if not os.path.exists(excel_folder_path):
            os.makedirs(excel_folder_path)

        excel_file_path = os.path.join(
            excel_folder_path, f"{unique_filename}.xlsx")
        wb.save(excel_file_path)

        file_format = request.args.get('file_format', 'excel')

        if file_format == 'pdf':
            pdf_folder_path = os.path.join(os.getcwd(), 'static')
            if not os.path.exists(pdf_folder_path):
                os.makedirs(pdf_folder_path)
            pdf_file_path = os.path.join(
                pdf_folder_path, f"{unique_filename}.pdf")
            pdf_file_path = excel_to_pdf(excel_file_path, pdf_file_path)
            # file_url = f"https://api.automhr.com/api/v1/download_report/{unique_filename}_filled.pdf"

            pdf_s3_key = f"pdfs/{unique_filename}.pdf"
            pdf_s3_url = upload_to_s3(pdf_file_path, pdf_s3_key)

            os.remove(excel_file_path)  # Delete the local Excel file
            os.remove(pdf_file_path)  # Delete the local PDF file

            file_url = pdf_s3_url

        else:
            # file_url = f"https://api.automhr.com/api/v1/download_report/{unique_filename}.xlsx"
            excel_s3_key = f"excels/{unique_filename}.xlsx"
            excel_s3_url = upload_to_s3(excel_file_path, excel_s3_key)

            os.remove(excel_file_path)  # Delete the local Excel file

            file_url = excel_s3_url

        return jsonify({"message": "Excel report generated successfully", "file_url": file_url})
    except Exception as e:
        logger.error("Error generating project report: %s", str(e))
        return jsonify({"message": "Error generating report", "error": str(e)})


@app.route('/api/v1/download_report/<filename>', methods=['GET'])
def download_timesheet_pdf(filename):
    try:
        static_dir = os.path.join(app.root_path, 'static')
        file_path = os.path.join(static_dir, filename)
        if os.path.isfile(file_path):
            return send_from_directory(static_dir, filename)
        else:
            return "File not found", 404
    except Exception as e:
        logger.error("Error downloading timesheet PDF: %s", str(e))
        return jsonify({'error': 'Failed to download timesheet PDF', 'message': str(e)}), 500


@app.route('/api/v1/test', methods=['GET'])
def test_api():
    return jsonify({"message": "Test Successfull"})


if __name__ == '__main__':
    app.run(debug=True)
